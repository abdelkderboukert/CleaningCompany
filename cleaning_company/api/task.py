from apscheduler.schedulers.background import BackgroundScheduler
from api.models import *
import datetime
import time
from openpyxl import Workbook
from django.db.models import Q
import os
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl
from datetime import datetime
from datetime import date

def my_task():
    print("Running my task!")
    Employees.objects.all().update(hourjob=0)

def export_to_excel():
    # Get the current date and time
    now = datetime.now()

    # Create a new Excel workbook
    wb = Workbook()

    # Get the active worksheet
    ws = wb.active

    # Set the header row
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = 'tablo'
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, color='FFFFFF', size=16)  # White bold text
    cell.fill = PatternFill(start_color='0000FF', fill_type='solid')  # Blue fill

    # Add borders to the header row
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    cell.border = border

    for col in range(1, 8):
        cell = ws.cell(row=2, column=col)
        cell.fill = PatternFill(start_color='FFFF00', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=14)
        cell.border = border  # Add border to each cell in the header row

    ws['A2'] = 'Employee ID'
    ws['B2'] = 'Name'
    ws['C2'] = 'Prename'
    ws['D2'] = 'Salary'
    ws['E2'] = 'Hour'
    ws['F2'] = 'Hour Job'
    ws['G2'] = 'Salary to pay'

    # Get the data from the database
    employees = Employees.objects.all()

    # Initialize total salary to pay
    total_salary_to_pay = 0

    # Iterate over the data and write it to the worksheet
    for i, employee in enumerate(employees, start=2):
        ws[f'A{i+1}'] = employee.id
        ws[f'B{i+1}'] = employee.name
        ws[f'C{i+1}'] = employee.prename
        ws[f'D{i+1}'] = employee.salary
        ws[f'E{i+1}'] = employee.hour
        ws[f'F{i+1}'] = employee.hourjob
        ws[f'G{i+1}'] = employee.salarypay
        total_salary_to_pay += employee.salarypay

        # Add borders to each cell in the data rows
        for col in range(1, 8):
            cell = ws.cell(row=i+1, column=col)
            cell.border = border

    total_row = len(employees) + 3    
    # Set the width of columns B to F to 90px
    for col in range(1, 8):  # Columns A to G
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True, color='FFFFFF', size=16)  # White bold text
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')  # Blue fill
        cell.border = border  # Add border to each cell in the total row
    
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 20

    # Add a row for the total salary to pay
    
    ws[f'A{total_row}'] = 'Total'
    ws.merge_cells(f'B{total_row}:G{total_row}')
    ws[f'B{total_row}'] = total_salary_to_pay

    # Specify the path where you want to save the file
    folder_path = r'C:\Users\HP\rebo\CleaningCompany\cleaning_company\ExcelFile'
    filename = f'{folder_path}/employees_{now.strftime("%Y-%m")}.xlsx'

    # Save the workbook to a file
    wb.save(filename)

    # Return the filename
    return filename

scheduler = BackgroundScheduler()
scheduler.add_job(my_task, 'cron', day='1', hour='0', minute='0')
scheduler.add_job(export_to_excel, 'cron', day='1', hour='0', minute='0')
scheduler.start()

# while True:
#     time.sleep(1)