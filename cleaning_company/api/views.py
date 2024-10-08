from django.shortcuts import render
from django.http import JsonResponse
from .serializers import *
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework import generics
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.request import Request
from django.db.models import F,Q
from datetime import date
from rest_framework import viewsets, filters
import os
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl
from datetime import datetime
from django.db.models import F
from django_filters.rest_framework import DjangoFilterBackend
from .models import *  # Import your Employees model

@api_view(['GET'])
def getRoutes(request):
    routes = [
        '/api/hourjob/',
        '/api/Employee/',
    ]
    return Response(routes)

class CreateEmployeeView(APIView):
    def post(self, request):
        serializer = EmployeeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)

class EmployeeView(APIView):
    def get(self, request):
        query = request.GET.get('q')
        if query:
            Employee = Employees.objects.filter(Q(name__icontains=query) | Q(prename__icontains=query))
        else:
            Employee = Employees.objects.all()
        serializer = EmployeesListeSerializer(Employee, many=True)
        return Response(serializer.data)
    
    def post(self, request):
        serializer = EmployeeSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        employee = Employees.objects.get(pk=pk)
        if employee:
            employee.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response(status=status.HTTP_404_NOT_FOUND)
    
class DeleteEmployeeView(APIView):
    def delete(self, request, pk):
        try:
            Employees.objects.get(id=pk).delete()
        except Employees.DoesNotExist:
            pass
        Response(status=status.HTTP_201_CREATED)

class HourJobView(APIView):
    def post(self, request):
        data = request.data
        errors = []
        for item in data.values():
            # Validate the item data
            if not all(key in item for key in ['date', 'employee', 'hours', 'notes']):
                errors.append({'error': 'Missing required keys', 'data': item})
                continue

            try:
                
                employee = Employees.objects.get(id=item['employee'])
                attendance = Attendance.objects.create(
                    employee=employee,
                    notes=item['notes'],
                    hours=item['hours'],
                    date=item['date']
                )
                attendance.save()
                employee.hourjob = employee.hourjob + int(item['hours'])
                employee.save()
                employee.salarypay = F('salary_per_hour') * F('hourjob') - F('accompte')
                employee.save()
            except Employees.DoesNotExist:
                errors.append({'error': f"Employee with id {item['employee']} does not exist", 'data': item})

        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        return Response(status=status.HTTP_202_ACCEPTED)

    def get(self, request):
        query = request.GET.get('q')
        if query:
            employees = Employees.objects.filter(name__icontains=query)
        else:
            employees = Employees.objects.all()

        # Update hourjob attribute in a single query
        employees.update(salarypay=F('salary_per_hour') * F('hourjob'))

        serializer = EmployeesListeSerializer(employees, many=True)
        return Response(serializer.data)
    
# class TodoView(APIView):
#     def get(self, request):
#         todo = Todo.objects.all()
#         serializer = TodoSerializer(todo, many=True)
#         return Response(serializer.data)
    
#     def post(self, request):
#         serializer = TodoSerializer(data=request.data, context={'request': request})
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
#     def patch(self, request, pk):
#         if pk is None:
#             # Handle the case where no pk is provided
#             return Response({'error': 'No pk provided'}, status=status.HTTP_400_BAD_REQUEST)
#         todo = Todo.objects.get(pk=pk)
#         serializer = TodoSerializer(todo, data=request.data, partial=True)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class TodoView(viewsets.ModelViewSet):
    queryset = Todo.objects.all()
    serializer_class = TodoSerializer

class AttendanceView(APIView):
    def get(self, request):
        query = request.GET.get('q')
        if query:
            Attendanc = Attendance.objects.filter(employee_id=int(query))
        else:
            Attendanc = Attendance.objects.all()
        serializer = AttendanceSerializer(Attendanc, many=True)
        return Response(serializer.data)

class TarifView(viewsets.ModelViewSet):
    queryset = Tarif.objects.all()
    serializer_class = TarifSerializer

def export_to_excel_view(request):
    # Get the current date and time
    now = datetime.now()
    print("is ok")
    # Create a new Excel workbook
    wb = Workbook()

    # Get the active worksheet
    ws = wb.active

    # Set the header row
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = 'tablo'
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, color='FFFFFF', size=16)  # White bold text
    cell.fill = PatternFill(start_color='0000FF', fill_type='solid')  # Blue fill

    # Add borders to the header row
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    cell.border = border

    for col in range(1, 8):
        cell = ws.cell(row=2, column=col)
        cell.fill = PatternFill(start_color='FFFF00', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=14)
        cell.border = border  # Add border to each cell in the header row

    ws['A2'] = 'Employee ID'
    ws['B2'] = 'Name'
    ws['C2'] = 'Prename'
    ws['D2'] = 'Salary'
    ws['E2'] = 'Hour'
    ws['F2'] = 'Hour Job'
    ws['G2'] = 'Salary to pay'

    # Get the data from the database
    employees = Employees.objects.all()

    # Initialize total salary to pay
    total_salary_to_pay = 0

    # Iterate over the data and write it to the worksheet
    for i, employee in enumerate(employees, start=2):
        ws[f'A{i+1}'] = employee.id
        ws[f'B{i+1}'] = employee.name
        ws[f'C{i+1}'] = employee.prename
        ws[f'D{i+1}'] = employee.salary
        ws[f'E{i+1}'] = employee.hour
        ws[f'F{i+1}'] = employee.hourjob
        ws[f'G{i+1}'] = employee.salarypay
        total_salary_to_pay += employee.salarypay

        # Add borders to each cell in the data rows
        for col in range(1, 8):
            cell = ws.cell(row=i+1, column=col)
            cell.border = border

    total_row = len(employees) + 3    
    # Set the width of columns B to F to 90px
    for col in range(1, 8):  # Columns A to G
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True, color='FFFFFF', size=16)  # White bold text
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')  # Blue fill
        cell.border = border  # Add border to each cell in the total row

    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 20

    # Add a row for the total salary to pay
    
    ws[f'A{total_row}'] = 'Total'
    ws.merge_cells(f'B{total_row}:G{total_row}')
    ws[f'B{total_row}'] = total_salary_to_pay

    # Create a file response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="employees_{now.strftime("%Y-%m")}.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    # Return the response
    return response


def export_to_excel_view2(request):
    # Get the current date and time
    now = datetime.now()

    # Create a new Excel workbook
    wb = Workbook()

    # Get the active worksheet
    ws = wb.active
    
    tabel = f'employees_{now.strftime("%Y-%m")}'

    # Set the header row
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = tabel
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, color='FFFFFF', size=16)  # White bold text
    cell.fill = PatternFill(start_color='0000FF', fill_type='solid')  # Blue fill

    # Add borders to the header row
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    cell.border = border

    for col in range(1, 5):
        cell = ws.cell(row=2, column=col)
        cell.fill = PatternFill(start_color='FFFF00', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=14)
        cell.border = border  # Add border to each cell in the header row

    ws['A2'] = 'tarif ID'
    ws['B2'] = 'item'
    ws['C2'] = 'prix'
    ws['D2'] = 'date'


    # Get the data from the database
    current_year = datetime.now().year
    current_month = datetime.now().month
    tarifs = Tarif.objects.filter(date__contains=f"{current_year}-{current_month:02}-")


    # Initialize total salary to pay
    total_tarif = 0

    # Iterate over the data and write it to the worksheet
    for i, tarif in enumerate(tarifs, start=2):
        ws[f'A{i+1}'] = tarif.id
        ws[f'B{i+1}'] = tarif.item
        ws[f'C{i+1}'] = tarif.monto
        ws[f'D{i+1}'] = tarif.date
        total_tarif += tarif.monto

        # Add borders to each cell in the data rows
        for col in range(1, 5):
            cell = ws.cell(row=i+1, column=col)
            cell.border = border

    total_row = len(tarifs) + 3    
    # Set the width of columns B to F to 90px
    for col in range(1, 5):  # Columns A to D
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True, color='FFFFFF', size=16)  # White bold text
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')  # Blue fill
        cell.border = border  # Add border to each cell in the total row
    
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 20

    # Add a row for the total salary to pay
    
    ws[f'A{total_row}'] = 'Total'
    ws.merge_cells(f'B{total_row}:D{total_row}')
    ws[f'B{total_row}'] = total_tarif

    # Create a file response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="tarif_{now.strftime("%Y-%m")}.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    # Return the response
    return response